import boto3
import pandas as pd
from openpyxl import load_workbook

# AWS S3 Configuration
s3_client = boto3.client('s3', region_name='us-west-1')

# Define bucket name and file paths
bucket_name = 'mzq-mzqplan-s3bucket'
file_key = 'mzqplan-aws-cloud/d_drive/wrapper_codes/Wrap Portal Login Generator_3.xlsm'
local_file_path = 'Wrap_Portal_Login_Generator_3.xlsm'

# Download the file from S3
s3_client.download_file(bucket_name, file_key, local_file_path)

# Load the Excel file using openpyxl
workbook = load_workbook(local_file_path, data_only=True)

# Print sheet names to verify
print("Sheet names:", workbook.sheetnames)

# Function to convert sheet to DataFrame
def sheet_to_df(sheet):
    data = sheet.values
    cols = next(data, None)
    if cols is None:
        print(f"No data in sheet: {sheet.title}")
        return pd.DataFrame()
    return pd.DataFrame(data, columns=cols)

# Check each sheet for data
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    df = sheet_to_df(sheet)
    
    if not df.empty:
        local_csv_path = f'{sheet_name}.csv'
        print(f"Data from sheet {sheet_name}:\n", df.head())
        df.to_csv(local_csv_path, index=False)
        print(f"Sheet {sheet_name} has been converted to {local_csv_path}")

        # Upload the CSV file back to S3
        csv_file_key = f'mzqplan-aws-cloud/d_drive/wrapper_codes/{sheet_name}.csv'
        s3_client.upload_file(local_csv_path, bucket_name, csv_file_key)
        print(f"CSV file {local_csv_path} has been uploaded to S3 at {csv_file_key}")
    else:
        print(f"Sheet {sheet_name} is empty or has no data rows.")

print("Conversion and upload process completed.")
