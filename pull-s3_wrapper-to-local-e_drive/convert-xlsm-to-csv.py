import boto3
import pandas as pd
from openpyxl import load_workbook
import os
import logging

# Setup logging
logging.basicConfig(level=logging.INFO)

# AWS S3 Configuration
s3_client = boto3.client('s3', region_name='us-west-1')

# Define bucket name and file paths
bucket_name = 'mzq-mzqplan-s3bucket'
file_key = 'mzqplan-aws-cloud/d_drive/wrapper_codes/Wrap Portal Login Generator_3.xlsm'
local_file_path = 'Wrap_Portal_Login_Generator_3.xlsm'

# Download the file from S3
logging.info(f"Downloading file {file_key} from S3 bucket {bucket_name}")
s3_client.download_file(bucket_name, file_key, local_file_path)

# Verify the file download
if os.path.exists(local_file_path) and os.path.getsize(local_file_path) > 0:
    logging.info(f"Successfully downloaded: {local_file_path}")
else:
    logging.error(f"File download failed or file is empty: {local_file_path}")
    exit(1)

# Load the Excel file using openpyxl
try:
    workbook = load_workbook(local_file_path, data_only=True)
    logging.info(f"Sheet names: {workbook.sheetnames}")
except Exception as e:
    logging.error(f"Failed to load workbook: {e}")
    exit(1)

# Function to convert sheet to DataFrame
def sheet_to_df(sheet):
    data = sheet.values
    cols = next(data, None)
    if cols is None:
        logging.warning(f"No data in sheet: {sheet.title}")
        return pd.DataFrame()
    return pd.DataFrame(data, columns=cols)

# Check each sheet for data
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    df = sheet_to_df(sheet)
    
    if not df.empty:
        local_csv_path = f'{sheet_name}.csv'
        logging.info(f"Data from sheet {sheet_name}:\n{df.head()}")
        df.to_csv(local_csv_path, index=False)
        logging.info(f"Sheet {sheet_name} has been converted to {local_csv_path}")

        # Upload the CSV file back to S3
        csv_file_key = f'mzqplan-aws-cloud/d_drive/wrapper_codes/{sheet_name}.csv'
        s3_client.upload_file(local_csv_path, bucket_name, csv_file_key)
        logging.info(f"CSV file {local_csv_path} has been uploaded to S3 at {csv_file_key}")

        # Optionally delete the local CSV file after upload
        if os.path.exists(local_csv_path):
            os.remove(local_csv_path)
            logging.info(f"Deleted local file: {local_csv_path}")
    else:
        logging.warning(f"Sheet {sheet_name} is empty or has no data rows.")

# Optionally delete the local Excel file after processing
if os.path.exists(local_file_path):
    os.remove(local_file_path)
    logging.info(f"Deleted local file: {local_file_path}")

logging.info("Conversion and upload process completed.")
